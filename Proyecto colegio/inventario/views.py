from django.shortcuts import render, redirect, get_object_or_404
from .models import Inventario
from .forms import InventarioForm
from django.http import HttpResponse , JsonResponse
import pandas as pd
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, authenticate
from io import BytesIO
from django.db.models import Q  
from django.db import models 
import barcode
from barcode.writer import ImageWriter
import base64
from django.views.decorators.csrf import csrf_exempt
from django.core.files.base import ContentFile
import random
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import os
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from io import BytesIO
import pandas as pd
from .models import Inventario
from PIL import Image as PILImage
from openpyxl.utils import get_column_letter
from django.conf import settings
from openpyxl.drawing.image import Image as OpenpyxlImage
from django.db.models.functions import Cast
from django.db.models import CharField




def index(request):
    return render(request, 'index.html')




def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            usuario = form.get_user()
            login(request, usuario)
            return redirect('inventario')  # ✅ Redirige a la vista del inventario
        else:
            return render(request, 'login.html', {
                'form': form,
                'hide_nav': True
            })
    else:
        form = AuthenticationForm()
        return render(request, 'login.html', {
            'form': form,
            'hide_nav': True
        })



from .models import Inventario

def inventario(request):
    palabra_clave = request.GET.get('q', '')
    codigo_barras = request.GET.get('codigo_barras', '')
    estado = request.GET.get('estado', '')
    ubicacion = request.GET.get('ubicacion', '')
    prestamo = request.GET.get('prestamo', '')  # 'on' si marcado

    equipos = Inventario.objects.all()

    if palabra_clave:
        equipos = equipos.filter(
            Q(equipo__icontains=palabra_clave) |
            Q(marca__icontains=palabra_clave) |
            Q(modelo__icontains=palabra_clave) |
            Q(detalle__icontains=palabra_clave) |
            Q(nombre_solicitante__icontains=palabra_clave)
        )

    if codigo_barras:
        equipos = equipos.filter(codigo__icontains=codigo_barras)

    if estado:
        equipos = equipos.filter(estado=estado)

    if ubicacion:
        equipos = equipos.filter(ubicacion=ubicacion)

    if prestamo == 'on':
        equipos = equipos.filter(nombre_solicitante__isnull=False)  # Solo los prestados

    # Ahora filtramos ESTADOS y UBICACIONES existentes en la BD, PERO verificamos que sean válidos
    estados_bd = Inventario.objects.values_list('estado', flat=True).distinct()
    ubicaciones_bd = Inventario.objects.values_list('ubicacion', flat=True).distinct()

    estados_validos = [choice[0] for choice in Inventario.ESTADO_CHOICES]
    ubicaciones_validas = [choice[0] for choice in Inventario.UBICACION_CHOICES]

    estados = [estado for estado in estados_bd if estado in estados_validos]
    ubicaciones = [ubicacion for ubicacion in ubicaciones_bd if ubicacion in ubicaciones_validas]

    context = {
        'equipos': equipos,
        'palabra_clave': palabra_clave,
        'codigo_barras': codigo_barras,
        'estado_seleccionado': estado,
        'ubicacion_seleccionada': ubicacion,
        'prestamo_seleccionado': prestamo,
        'estados': estados,
        'ubicaciones': ubicaciones,
    }
    return render(request, 'inventario.html', context)





def inventario_view(request):
    equipos = Inventario.objects.all()
    palabra_clave = request.GET.get('q', '')
    if palabra_clave:
        equipos = equipos.filter(
            Q(equipo__icontains=palabra_clave) |
            Q(tipo_equipo__icontains=palabra_clave) |
            Q(marca__icontains=palabra_clave) |
            Q(modelo__icontains=palabra_clave) |
            Q(ubicacion__icontains=palabra_clave)
        )

    codigo_barras = request.GET.get('codigo_barras', '')
    if codigo_barras:
        equipos = equipos.filter(codigo__icontains=codigo_barras)

    equipos = equipos.order_by('equipo')  # ✅ orden alfabético

    return render(request, 'inventario.html', {
        'equipos': equipos,
        'palabra_clave': palabra_clave,
        'codigo_barras': codigo_barras
    })



def registrar_equipo(request):
    if request.method == 'POST':
        form = InventarioForm(request.POST, request.FILES)
        if form.is_valid():
            equipo = form.save()  # El modelo genera código + código de barras
            return redirect('inventario')
    else:
        form = InventarioForm()
    return render(request, 'registrar_equipo.html', {'form': form})


@csrf_exempt
def generar_codigo_barra(request):
    codigo = request.GET.get('codigo', '')
    if codigo:
        try:
            if len(codigo) == 12:
                EAN13 = barcode.get_barcode_class('ean13')
                barcode_obj = EAN13(codigo, writer=ImageWriter())
            else:
                CODE128 = barcode.get_barcode_class('code128')
                barcode_obj = CODE128(codigo, writer=ImageWriter())
            buffer = BytesIO()
            barcode_obj.write(buffer)
            img_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
            return JsonResponse({'barcode': img_data})
        except Exception as e:
            return JsonResponse({'error': f"Error generando código de barras: {str(e)}"}, status=500)
    return JsonResponse({'error': 'No se proporcionó código'}, status=400)

# (Resto de tus vistas: exportar_excel, escanear_codigo, agregar_inventario, actualizar_prestamo, buscar_inventario, editar_equipo, eliminar_equipo, detalle_equipo)



def listar_inventario(request):
    inventarios = Inventario.objects.all()
    return render(request, 'inventario.html', {'inventarios': inventarios})




from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.conf import settings
import os
from .models import Inventario

def exportar_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = [
        "Equipo", "Tipo", "Marca", "Modelo", "Ubicación", "Estado", "Detalle", "Código",
        "Fecha Adquisición", "Fondo Adquisición", "Préstamo Activo", "Nombre Solicitante",
        "Fecha Inicio Préstamo", "Fecha Fin Préstamo", "Duración (días)", "Foto", "Código de Barras"
    ]
    ws.append(headers)

    # 🎯 Capturamos filtros como en inventario
    palabra_clave = request.GET.get('q', '')
    codigo_barras = request.GET.get('codigo_barras', '')
    estado = request.GET.get('estado', '')
    ubicacion = request.GET.get('ubicacion', '')
    prestamo = request.GET.get('prestamo', '')

    inventarios = Inventario.objects.all()

    if palabra_clave:
        inventarios = inventarios.filter(
            Q(equipo__icontains=palabra_clave) |
            Q(marca__icontains=palabra_clave) |
            Q(modelo__icontains=palabra_clave) |
            Q(detalle__icontains=palabra_clave) |
            Q(nombre_solicitante__icontains=palabra_clave)
        )

    if codigo_barras:
        inventarios = inventarios.filter(codigo__icontains=codigo_barras)

    if estado:
        inventarios = inventarios.filter(estado=estado)

    if ubicacion:
        inventarios = inventarios.filter(ubicacion=ubicacion)

    if prestamo == 'on':
        inventarios = inventarios.filter(nombre_solicitante__isnull=False)

    # Ahora exportamos SOLO lo filtrado
    for row_num, inv in enumerate(inventarios, start=2):
        ws.cell(row=row_num, column=1, value=inv.equipo)
        ws.cell(row=row_num, column=2, value=inv.tipo_equipo)
        ws.cell(row=row_num, column=3, value=inv.marca)
        ws.cell(row=row_num, column=4, value=inv.modelo)
        ws.cell(row=row_num, column=5, value=inv.ubicacion)
        ws.cell(row=row_num, column=6, value=inv.estado)
        ws.cell(row=row_num, column=7, value=inv.detalle)
        ws.cell(row=row_num, column=8, value=inv.codigo)
        ws.cell(row=row_num, column=9, value=inv.fecha_adquisicion.strftime('%d/%m/%Y') if inv.fecha_adquisicion else "-")
        ws.cell(row=row_num, column=10, value=inv.fondo_adquisicion or "-")
        ws.cell(row=row_num, column=11, value="Sí" if inv.prestamo_activo else "No")
        ws.cell(row=row_num, column=12, value=inv.nombre_solicitante or "-")
        ws.cell(row=row_num, column=13, value=inv.fecha_inicio_prestamo.strftime('%d/%m/%Y') if inv.fecha_inicio_prestamo else "-")
        ws.cell(row=row_num, column=14, value=inv.fecha_fin_prestamo.strftime('%d/%m/%Y') if inv.fecha_fin_prestamo else "-")
        
        duracion = (inv.fecha_fin_prestamo - inv.fecha_inicio_prestamo).days if inv.fecha_inicio_prestamo and inv.fecha_fin_prestamo else "-"
        ws.cell(row=row_num, column=15, value=f"{duracion} días" if isinstance(duracion, int) else "-")

        # FOTO
        if inv.foto:
            foto_path = os.path.join(settings.MEDIA_ROOT, str(inv.foto))
            if os.path.exists(foto_path):
                img_foto = OpenpyxlImage(foto_path)
                img_foto.width = 130
                img_foto.height = 75
                ws.column_dimensions['P'].width = 23
                ws.add_image(img_foto, f"P{row_num}")
            else:
                ws.cell(row=row_num, column=16, value="Foto no disponible")
        else:
            ws.cell(row=row_num, column=16, value="Foto no disponible")

        # CÓDIGO DE BARRAS
        if inv.codigo_barras:
            codigo_path = os.path.join(settings.MEDIA_ROOT, str(inv.codigo_barras))
            if os.path.exists(codigo_path):
                img_cod = OpenpyxlImage(codigo_path)
                img_cod.width = 130
                img_cod.height = 75
                ws.column_dimensions['Q'].width = 23
                ws.add_image(img_cod, f"Q{row_num}")
            else:
                ws.cell(row=row_num, column=17, value="No generado")
        else:
            ws.cell(row=row_num, column=17, value="No generado")

        ws.row_dimensions[row_num].height = 80

    # Ajuste compacto de columnas de texto
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        if col_letter not in ["P", "Q"]:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            ws.column_dimensions[col_letter].width = min(max_length * 1.1, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventario.xlsx"'
    return response









def escanear_codigo(request):
    return render(request, 'inventario.html')


def agregar_inventario(request):
    if request.method == 'POST':
        form = InventarioForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()  # aquí se calculan los días (opción B) o en el modelo (opción A).
            return redirect('inventario')
    else:
        form = InventarioForm()
    return render(request, 'inventario.html', {'form': form})


def actualizar_prestamo(request, id):
    inventario = get_object_or_404(Inventario, id=id)
    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            form.save()
            messages.success(request, "Estado actualizado correctamente.")
            return redirect('listar_inventario')
    else:
        form = InventarioForm(instance=inventario)
    return render(request, 'actualizar_prestamo.html', {'form': form, 'inventario': inventario})


from django.db.models import Q

from django.db.models import Q
from django.utils.dateformat import DateFormat
from .models import Inventario

def buscar_inventario(request):
    query = request.GET.get('q', '').strip()
    codigo_barras = request.GET.get('codigo_barras', '').strip()

    inventarios = Inventario.objects.all()

    if query:
        # Para fechas, convertimos a texto
        inventarios = inventarios.annotate(
            fecha_adq_str=models.functions.Cast('fecha_adquisicion', models.CharField()),
            fecha_inicio_str=models.functions.Cast('fecha_inicio_prestamo', models.CharField()),
            fecha_fin_str=models.functions.Cast('fecha_fin_prestamo', models.CharField())
        ).filter(
            Q(equipo__icontains=query) |
            Q(tipo_equipo__icontains=query) |
            Q(marca__icontains=query) |
            Q(modelo__icontains=query) |
            Q(ubicacion__icontains=query) |
            Q(estado__icontains=query.replace(" ", "_").lower()) |
            Q(codigo__icontains=query) |
            Q(foto__icontains=query) |
            Q(fondo_adquisicion__icontains=query) |
            Q(nombre_solicitante__icontains=query) |
            Q(fecha_adq_str__icontains=query) |
            Q(fecha_inicio_str__icontains=query) |
            Q(fecha_fin_str__icontains=query)
        )

    if codigo_barras:
        inventarios = inventarios.filter(codigo__icontains=codigo_barras)

    return render(request, 'inventario.html', {
        'equipos': inventarios,
        'palabra_clave': query,
        'codigo_barras': codigo_barras
    })


def editar_equipo(request, equipo_id):
    equipo = get_object_or_404(Inventario, id=equipo_id)
    if request.method == 'POST':
        form = InventarioForm(request.POST, request.FILES, instance=equipo)
        if form.is_valid():
            equipo = form.save(commit=False)
            fecha_adquisicion = request.POST.get('fecha_adquisicion')
            if fecha_adquisicion:
                equipo.fecha_adquisicion = fecha_adquisicion
            equipo.save()
            return redirect('inventario')
    else:
        form = InventarioForm(instance=equipo)
    return render(request, 'editar_equipo.html', {'form': form, 'equipo': equipo})




def eliminar_equipo(request, equipo_id):
    equipo = get_object_or_404(Inventario, id=equipo_id)
    if request.method == 'POST':
        equipo.delete()
        return redirect('inventario')
    return redirect('inventario')

def registrar_equipo(request):
    if request.method == 'POST':
        form = InventarioForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('inventario')  # Redirige a Inventario en lugar de editar_equipo
    else:
        form = InventarioForm()
    return render(request, 'registrar_equipo.html', {'form': form})



def detalle_equipo(request, id):
    equipo = get_object_or_404(Inventario, id=id)
    return render(request, 'detalle_equipo.html', {'equipo': equipo})


def __str__(self):
        return f"{self.equipo} - {self.marca} {self.modelo}"